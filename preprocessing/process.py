"""
preprocess.py — Improved Grade Worksheet Preprocessor
======================================================
Reads the "Consolidated Worksheet of Grades" Excel format:
  - Skips header block (university name, title, college name, semester)
  - Detects course name rows (e.g. "Bachelor of Science in Architecture")
  - Handles the alternating subject-header / grade-data row pattern
  - Captures Student_ID, Gender, Status, Course_Subject_Name, Grade, GWA,
    Semester, College, Course, Year
  - Handles INC/X.XX and DROP/DRP grade notations
  - Outputs:
      1. Final_Merged_Student_Data.csv          (all models base)
      2. model_datasets/<model>_data.csv        (one per model)

Usage:
    python preprocess.py --files path1.xlsx path2.xlsx ... [--year YEAR]
    python preprocess.py                        # uses FILES dict below
"""

import pandas as pd
import numpy as np
import os
import re
import argparse
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────
# CONFIG — edit these or pass via CLI
# ─────────────────────────────────────────────────────────────

# Default file list when running without CLI args
FILES = {
    "2022-2023": "datasets/2022-1_Consolidated_Worksheet_of_Grades.xlsx",
}

PROCESSED_DIR  = "processed_datasets"
MODEL_DATA_DIR = os.path.join(PROCESSED_DIR, "model_datasets")
FINAL_OUTPUT   = os.path.join(PROCESSED_DIR, "Final_Merged_Student_Data.csv")

# Sentinel values stored in Grade column
GRADE_INC  = 5.0   # INC with no revealed grade
GRADE_DROP = 0.0   # Dropped subject

# Columns in the final merged file
FINAL_COLUMNS = [
    "Student_ID", "Gender", "Status",
    "Course_Subject_Name", "Grade", "GWA",
    "Semester", "College", "Course", "Year",
]


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def _is_course_row(row) -> bool:
    """Row[0] is a non-empty string containing a degree keyword."""
    v = row[0]
    if not isinstance(v, str):
        return False
    v_up = v.strip().upper()
    keywords = ["BACHELOR", "MASTER", "DOCTOR", "DIPLOMA",
                "ASSOCIATE", "CERTIFICATE", "BS ", "AB "]
    return any(k in v_up for k in keywords)


def _is_subject_header_row(row) -> bool:
    """Row[3..] has subject codes (alphanumeric 6-10 chars), row[0] is None/empty."""
    if row[0] is not None and str(row[0]).strip() != "":
        return False
    codes = [c for c in row[3:] if c and isinstance(c, str)
             and re.match(r'^[A-Z]{2,6}[\d\-\/]{2,}', str(c).strip())]
    return len(codes) >= 1


def _is_data_row(row) -> bool:
    """Row[0] is an integer (sequential student number)."""
    return isinstance(row[0], (int, float)) and row[0] == int(row[0]) and row[0] >= 1


def _parse_grade(val):
    """
    Parse a grade cell value.
    Returns (grade_float, status_str)
    Handles: 1.75, "2.25", "INC/5.00", "INC", "DROP", "5.00", None
    """
    if val is None:
        return None, "REGULAR"

    s = str(val).strip().upper()

    # Empty
    if s in ("", "NONE", "-", "N/A"):
        return None, "REGULAR"

    # INC pattern: "INC", "INC/5.00", "INC/2.50"
    if "INC" in s:
        match = re.search(r'[\d]+\.[\d]+', s)
        grade = float(match.group()) if match else GRADE_INC
        return grade, "INC"

    # DROP / DRP pattern
    if "DROP" in s or "DRP" in s:
        return GRADE_DROP, "DROP"

    # Numeric grade
    try:
        g = float(s)
        # 5.0 often means failed (failing mark in PH grading)
        return g, "REGULAR"
    except ValueError:
        return None, "REGULAR"


def _extract_semester_year(text: str):
    """
    Parse '1st Semester, School Year 2022-2023' →
    ('1sem', '2022-2023')
    """
    sem = "1sem"
    year = "2022-2023"

    if not text:
        return sem, year

    t = str(text).upper()

    if "2ND" in t or "SECOND" in t:
        sem = "2sem"
    elif "SUMMER" in t:
        sem = "Summer"
    else:
        sem = "1sem"

    ym = re.search(r'(\d{4})[–\-](\d{4})', text)
    if ym:
        year = f"{ym.group(1)}-{ym.group(2)}"

    return sem, year


# ─────────────────────────────────────────────────────────────
# SHEET PARSER
# ─────────────────────────────────────────────────────────────

def parse_sheet(ws, sheet_name: str, default_year: str) -> pd.DataFrame:
    """
    Parse one worksheet from the consolidated grades workbook.
    Returns a long-form DataFrame.
    """
    rows = list(ws.iter_rows(values_only=True))

    # ── 1. Extract header metadata ──────────────────────────
    college  = sheet_name.strip().upper()
    semester = "1sem"
    year     = default_year

    for row in rows[:20]:
        flat = " ".join(str(v) for v in row if v)
        # College name
        col_match = re.search(
            r'(COLLEGE OF [A-Z ,&]+|DEPARTMENT OF [A-Z ,&]+)', flat
        )
        if col_match:
            college = col_match.group(1).strip()

        # Semester + Year
        sem_match = re.search(r'(\d(?:st|nd|rd|th)?)\s*(?:Semester|Sem)', flat, re.I)
        yr_match  = re.search(r'(\d{4})[–\-](\d{4})', flat)
        summer    = "SUMMER" in flat.upper()

        if sem_match or summer or yr_match:
            if summer:
                semester = "Summer"
            elif sem_match:
                n = sem_match.group(1)[0]
                semester = f"{n}sem"
            if yr_match:
                year = f"{yr_match.group(1)}-{yr_match.group(2)}"

    # ── 2. Walk rows and build records ───────────────────────
    records = []
    current_course  = "Unknown"
    current_subjects = []          # list of subject codes (from header row)

    i = 0
    while i < len(rows):
        row = rows[i]

        # Course name row
        if _is_course_row(row):
            current_course = str(row[0]).strip()
            current_subjects = []  # reset for new course section
            i += 1
            continue

        # Subject header row
        if _is_subject_header_row(row):
            # Subjects start at col 3
            current_subjects = []
            for v in row[3:]:
                if v and isinstance(v, str) and re.match(r'^[A-Z]{2,6}[\d\-\/]', str(v).strip()):
                    current_subjects.append(str(v).strip())
                elif v is None:
                    current_subjects.append(None)  # preserve position
                else:
                    current_subjects.append(None)
            i += 1
            continue

        # Data row (student grades)
        if _is_data_row(row):
            seq_id = int(row[0])
            gender_raw = str(row[1]).strip().capitalize() if row[1] else "Unknown"
            gender = 0 if gender_raw.lower() == "male" else (1 if gender_raw.lower() == "female" else -1)

            # GWA is usually last numeric in range 1.0-5.0 on the right side
            gwa = None
            for v in reversed(row[15:]):
                try:
                    candidate = float(v)
                    if 1.0 <= candidate <= 5.0:
                        gwa = candidate
                        break
                except (TypeError, ValueError):
                    pass

            grade_vals = list(row[3: 3 + len(current_subjects)]) if current_subjects else []

            student_uid = f"{seq_id}_{college[:6]}_{current_course[:10]}_{semester}_{year}"

            for j, subj in enumerate(current_subjects):
                if subj is None:
                    continue
                raw_val = grade_vals[j] if j < len(grade_vals) else None
                grade, subject_status = _parse_grade(raw_val)

                records.append({
                    "Student_ID"          : student_uid,
                    "Gender"              : gender,
                    "Course_Subject_Name" : subj,
                    "Grade"               : grade if grade is not None else np.nan,
                    "Subject_Status"      : subject_status,
                    "GWA"                 : gwa if gwa is not None else np.nan,
                    "Semester"            : semester,
                    "College"             : college,
                    "Course"              : current_course,
                    "Year"                : year,
                })

            i += 1
            continue

        i += 1

    if not records:
        return pd.DataFrame()

    df = pd.DataFrame(records)

    # Ensure Status col exists
    if "Status" not in df.columns:
        df["Status"] = "REGULAR"

    # Resolve student-level status from subject-level statuses
    if "Subject_Status" in df.columns:
        status_map = (
            df.groupby("Student_ID")["Subject_Status"]
            .apply(lambda x: "DROP" if "DROP" in x.values else ("INC" if "INC" in x.values else "REGULAR"))
        )
        df["Status"] = df["Student_ID"].map(status_map)
        df = df.drop(columns=["Subject_Status"])

    # Drop rows where Grade is still NaN
    df = df.dropna(subset=["Grade"]).reset_index(drop=True)

    # Round GWA
    df["GWA"] = pd.to_numeric(df["GWA"], errors="coerce").round(2)

    # Fix GWA=NaN/0 for INC students by computing from real grades
    real_grades = df[~df["Grade"].isin([GRADE_DROP, GRADE_INC])].groupby("Student_ID")["Grade"].mean()
    needs_fix   = df.groupby("Student_ID")["GWA"].first().pipe(lambda s: s[s.isna() | (s == 0)]).index
    fix_map     = real_grades.loc[real_grades.index.intersection(needs_fix)].round(2)
    df.loc[df["Student_ID"].isin(fix_map.index), "GWA"] = df["Student_ID"].map(fix_map)

    # Keep only final columns that exist
    keep = [c for c in FINAL_COLUMNS if c in df.columns]
    return df[keep].reset_index(drop=True)


# ─────────────────────────────────────────────────────────────
# FILE PROCESSOR
# ─────────────────────────────────────────────────────────────

def process_file(filepath: str, year: str) -> pd.DataFrame:
    print(f"\n{'='*60}")
    print(f"  File : {filepath}")
    print(f"  Year : {year}")
    print(f"{'='*60}")

    wb = load_workbook(filepath, read_only=True, data_only=True)
    frames = []

    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n  → Sheet: '{name}'")
        try:
            df = parse_sheet(ws, name, year)
            if df.empty:
                print(f"    [SKIP] No data parsed")
            else:
                frames.append(df)
                print(f"    [OK]   {len(df):,} rows | "
                      f"Students: {df['Student_ID'].nunique()} | "
                      f"Courses: {df['Course'].nunique() if 'Course' in df.columns else '?'}")
        except Exception as e:
            import traceback
            print(f"    [ERROR] {e}")
            traceback.print_exc()

    wb.close()
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# ─────────────────────────────────────────────────────────────
# MODEL-SPECIFIC DATASET EXPORTERS
# ─────────────────────────────────────────────────────────────

def export_model_datasets(df: pd.DataFrame, out_dir: str):
    """
    Produce a purpose-built CSV for each ML model.
    Each file contains exactly the columns that model needs,
    with model-specific cleaning / feature engineering baked in.
    """
    os.makedirs(out_dir, exist_ok=True)

    # ── Helper ───────────────────────────────────────────
    def save(name, frame, note=""):
        path = os.path.join(out_dir, f"{name}.csv")
        frame.to_csv(path, index=False)
        print(f"  [{name}]  {len(frame):,} rows  →  {path}  {note}")

    print(f"\n{'─'*60}")
    print("  Exporting model-specific datasets …")
    print(f"{'─'*60}")

    # ── 1. Dropout Risk per Student ─────────────────────
    # Used by: F-drop_outrisk_percollege_.py
    student_agg = df.groupby("Student_ID").agg(
        Gender   = ("Gender",  "first"),
        College  = ("College", "first"),
        Course   = ("Course",  "first") if "Course" in df.columns else ("College", "first"),
        Semester = ("Semester","first"),
        Year     = ("Year",    "first"),
        Avg_Grade= ("Grade",   "mean"),
        Sub_Count= ("Grade",   "count"),
        GWA      = ("GWA",     "first"),
    ).reset_index()

    student_agg["is_drop"]       = df.groupby("Student_ID")["Grade"].apply(lambda x: int((x == GRADE_DROP).any())).values
    student_agg["is_inc"]        = df.groupby("Student_ID")["Grade"].apply(lambda x: int((x == GRADE_INC).any())).values
    student_agg["fail_rate"]     = df.groupby("Student_ID")["Grade"].apply(lambda x: (x >= 3.0).mean()).values
    student_agg["Year_Numeric"]  = student_agg["Year"].str.extract(r"(\d{4})")[0].astype(float)
    save("dropout_risk_per_student", student_agg,
         f"| Students: {len(student_agg)}")

    # ── 2. Dropout Spike (cohort-level rates) ───────────
    # Used by: train_drop_spike.py
    cohort = student_agg.groupby(["Year_Numeric", "College"]).agg(
        Total_Students = ("Student_ID", "count"),
        Dropout_Count  = ("is_drop",    "sum"),
    ).reset_index()
    cohort["Dropout_Rate"] = (cohort["Dropout_Count"] / cohort["Total_Students"] * 100).round(2)
    cohort = cohort[cohort["Total_Students"] > 5]
    save("dropout_spike", cohort)

    # ── 3. Dropout Ranking per College ──────────────────
    # Used by: train_dropRanking_perCollege.py
    rank_df = student_agg[["Student_ID", "College", "Course", "Semester", "Year_Numeric", "is_drop"]].copy()
    save("dropout_ranking_per_college", rank_df)

    # ── 4. GWA Ranking per College ───────────────────────
    # Used by: train_gwaRanking_perCollege.py
    valid_gwa = df[(df["GWA"] >= 1.0) & (df["GWA"] <= 5.0)].copy()
    valid_gwa["Year_Numeric"] = valid_gwa["Year"].str.extract(r"(\d{4})")[0].astype(float)
    sem_map = {"1sem": 1, "2sem": 2, "summer": 3}
    valid_gwa["Sem_Numeric"] = valid_gwa["Semester"].str.lower().map(sem_map).fillna(1)
    gwa_rank = valid_gwa[["College", "Course", "Year_Numeric", "Sem_Numeric", "GWA"]].drop_duplicates()
    save("gwa_ranking_per_college", gwa_rank,
         f"| Valid GWA rows: {len(gwa_rank)}")

    # ── 5. GWA Trend ─────────────────────────────────────
    # Used by: train_gwatrend.py
    trend = valid_gwa[["College", "Course", "Year_Numeric", "Sem_Numeric", "GWA"]].copy()
    save("gwa_trend", trend)

    # ── 6. INC Forecast ──────────────────────────────────
    # Used by: train_inc_forecast.py
    student_agg["has_inc"] = student_agg["is_inc"]
    inc_cohort = student_agg.groupby(["Year_Numeric", "College"]).agg(
        Total_Students  = ("Student_ID", "count"),
        INC_Count       = ("has_inc",    "sum"),
    ).reset_index()
    inc_cohort["INC_Rate"] = (inc_cohort["INC_Count"] / inc_cohort["Total_Students"] * 100).round(2)
    inc_cohort = inc_cohort[inc_cohort["Total_Students"] > 5]
    save("inc_forecast", inc_cohort)

    # ── 7. Irregular vs Regular ──────────────────────────
    # Used by: train_irreg-reg.py
    student_agg["is_irregular"] = ((student_agg["is_drop"] == 1) | (student_agg["is_inc"] == 1)).astype(int)
    student_agg["Year_Numeric"] = student_agg["Year_Numeric"].astype(float)
    sem_map2 = {"1sem": 1, "2sem": 2, "summer": 3}
    student_agg["Sem_Numeric"] = student_agg["Semester"].str.lower().map(sem_map2).fillna(1)

    irreg_cohort = student_agg.groupby(["Year_Numeric", "College", "Sem_Numeric"]).agg(
        Total_Students  = ("Student_ID", "count"),
        Irregular_Count = ("is_irregular","sum"),
    ).reset_index()
    irreg_cohort["Irregular_Rate"] = (irreg_cohort["Irregular_Count"] / irreg_cohort["Total_Students"] * 100).round(2)
    irreg_cohort = irreg_cohort[irreg_cohort["Total_Students"] > 5]
    save("irreg_reg_status", irreg_cohort)

    # ── 8. KPI Students (GWA + Enrollment) ──────────────
    # Used by: train_KPI_students.py
    kpi_gwa = valid_gwa[["College", "Course", "Year_Numeric", "Sem_Numeric", "GWA"]].copy()
    save("kpi_gwa", kpi_gwa)

    enroll_df = student_agg.groupby(["Year_Numeric", "College"])["Student_ID"].nunique().reset_index()
    enroll_df.columns = ["Year_Numeric", "College", "Count"]
    save("kpi_enrollment", enroll_df)

    # ── 9. Subject Grade Forecast ────────────────────────
    # Used by: train_subject_top.py
    subj_df = df.copy()
    subj_df["Year_Numeric"] = subj_df["Year"].str.extract(r"(\d{4})")[0].astype(float)
    subj_df = subj_df[(subj_df["Grade"] >= 1.0) & (subj_df["Grade"] <= 5.0)]
    subj_df["Subject"] = subj_df["Course_Subject_Name"].str.upper().str.strip()
    subj_trend = subj_df.groupby(["Year_Numeric", "College", "Course", "Subject"])["Grade"].mean().reset_index()
    save("subject_grade_forecast", subj_trend,
         f"| Subjects: {subj_trend['Subject'].nunique()}")

    print(f"\n  ✓ All model datasets saved to: {out_dir}")


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Grade worksheet preprocessor")
    parser.add_argument("--files", nargs="*",
                        help="Excel files to process (e.g. 2022_sem1.xlsx 2023_sem2.xlsx)")
    parser.add_argument("--year", default=None,
                        help="Academic year override, e.g. 2022-2023")
    args = parser.parse_args()

    os.makedirs(PROCESSED_DIR,  exist_ok=True)
    os.makedirs(MODEL_DATA_DIR, exist_ok=True)

    # Build file → year mapping
    if args.files:
        file_map = {}
        for f in args.files:
            # Try to extract year from filename, else use --year or default
            ym = re.search(r'(\d{4})[_\-](\d{4})', os.path.basename(f))
            if ym:
                y = f"{ym.group(1)}-{ym.group(2)}"
            else:
                y = args.year or "2022-2023"
            file_map[y] = f
    else:
        file_map = FILES

    all_dfs = []
    for year, path in file_map.items():
        if not os.path.exists(path):
            print(f"\n[MISSING] {path}  — skipping")
            continue

        df = process_file(path, year)
        if df.empty:
            print(f"  [WARN] No data extracted from {path}")
            continue

        # Per-file CSV (in processed_datasets root)
        label = year.replace("-", "_")
        out   = os.path.join(PROCESSED_DIR, f"{label}_cleaned.csv")
        df.to_csv(out, index=False)
        print(f"\n  Saved per-file CSV → {out}  ({len(df):,} rows)")
        all_dfs.append(df)

    if not all_dfs:
        print("\n[ERROR] No data to process. Exiting.")
        return

    # Merge all years
    final = pd.concat(all_dfs, ignore_index=True)
    keep  = [c for c in FINAL_COLUMNS if c in final.columns]
    final = final[keep]
    if "GWA" in final.columns:
        final["GWA"] = final["GWA"].round(2)

    final.to_csv(FINAL_OUTPUT, index=False)

    print(f"\n{'='*60}")
    print(f"FINAL DATASET  →  {FINAL_OUTPUT}")
    print(f"  Rows       : {len(final):,}")
    print(f"  Students   : {final['Student_ID'].nunique():,}")
    if "College" in final.columns:
        print(f"  Colleges   : {final['College'].nunique()}")
    if "Course" in final.columns:
        print(f"  Courses    : {final['Course'].nunique()}")
    if "Status" in final.columns:
        print(f"\nStatus breakdown:")
        print(final["Status"].value_counts().to_string())

    # Export per-model datasets
    export_model_datasets(final, MODEL_DATA_DIR)


if __name__ == "__main__":
    main()