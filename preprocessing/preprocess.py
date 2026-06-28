"""
preprocess.py  –  NovaSight Grade-Sheet Preprocessor (v2)
==========================================================
Fixes vs the original:
  1. Parses EVERY sheet in every workbook (previously missed CCST, CNM, etc.)
  2. Detects ALL course blocks per sheet (previously only grabbed the first)
  3. Generates one long-form row per student × subject (~460k rows total)
  4. Correctly maps sheet codes → full college names
  5. Handles all special grade values: INC, INC/x.xx, DRP, NGA, NGA/x.xx, 0, 5
  6. Extracts semester & academic year from the sheet header text
  7. Builds all 12 model-dataset CSVs from the long-form master
  8. Can be called standalone OR imported by upload_routes.py

Usage (standalone):
    python preprocess.py  --input_dir Unprocessed_Datasets \
                          --output_dir Processed_Datasets
"""

import os
import re
import sys
import argparse
import logging
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Sheet code → unified college label ───────────────────────────────────────
# Grouping rules (from institution):
#   CEA  / COEA                              → CEA
#   CTEC                                     → CTEC
#   CCST                                     → CCST
#   COAS / DOAS                              → COAS
#   CNM  / CAHS-SOM / CAHS-SON / CAHS-SPHCD → CAHS
#   COBA / CBA                               → CBA
SHEET_COLLEGE_MAP = {
    # Engineering & Architecture (CEA and COEA are the same college)
    "CEA":        "CEA",
    "COEA":       "CEA",
    # Technology
    "CTEC":       "CTEC",
    # Computer Studies
    "CCST":       "CCST",
    # Arts & Sciences (DOAS renamed to COAS)
    "COAS":       "COAS",
    "DOAS":       "COAS",
    # Health Sciences (all CAHS sub-units grouped under one label)
    "CNM":        "CAHS",
    "CAHS-SOM":   "CAHS",
    "CAHS-SON":   "CAHS",
    "CAHS-SPHCD": "CAHS",
    # Business & Accountancy (COBA renamed to CBA)
    "COBA":       "CBA",
    "CBA":        "CBA",
}

# ── Dashboard short-code → CSV college labels ─────────────────────────────────
# Used by ml_analysis.py expand_college(). Keep in sync with SHEET_COLLEGE_MAP.
COLLEGE_MAP = {
    "CEA":  ["CEA"],
    "CTEC": ["CTEC"],
    "CCST": ["CCST"],
    "COAS": ["COAS"],
    "CAHS": ["CAHS"],
    "CBA":  ["CBA"],
}

# Semester text patterns inside the sheet header
SEM_PATTERNS = {
    re.compile(r"1st\s+semester", re.I): "1sem",
    re.compile(r"first\s+semester", re.I): "1sem",
    re.compile(r"2nd\s+semester", re.I): "2sem",
    re.compile(r"second\s+semester", re.I): "2sem",
    re.compile(r"summer", re.I): "summer",
}

YEAR_PATTERN = re.compile(r"(\d{4})\s*[-–]\s*(\d{4})")

# Grade special value encoding
GRADE_ENCODING = {
    "DRP":  0.0,   # Dropped
    "NGA":  0.0,   # No Grade (treated as drop)
    "INC":  5.0,   # Incomplete (worst grade bucket)
    "W":    0.0,   # Withdrawn
}


# ══════════════════════════════════════════════════════════════════════════════
#  PARSING HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_grade(raw) -> float | None:
    """
    Convert any raw cell value to a float grade or None (skip).

    Handles:
      - Standard numeric grades: 1.0, 1.25 … 3.0, 5.0
      - Dropped / no-grade: 0, DRP, NGA  → 0.0
      - Incomplete: INC                  → 5.0
      - Combined: INC/2.75, NGA/5.00     → 5.0 / 0.0 (prefix wins)
      - GWA summary cells (large floats or int > 5) → None
    """
    if raw is None:
        return None

    s = str(raw).strip()
    if s == "" or s.lower() == "none":
        return None

    # Combined grade like "INC/2.75" or "NGA/5.00"
    if "/" in s:
        prefix = s.split("/")[0].strip().upper()
        return GRADE_ENCODING.get(prefix, None)

    # Pure keyword
    upper = s.upper()
    if upper in GRADE_ENCODING:
        return GRADE_ENCODING[upper]

    # Numeric
    try:
        f = float(s)
        # Filter out summary/GWA cells: valid subject grades are 0–5
        if f < 0 or f > 5.0:
            return None
        return round(f, 4)
    except ValueError:
        return None


def is_subject_code(v) -> bool:
    """
    True if the value looks like a subject code (e.g. EGEC0103, MFHC0111-P).
    Subject codes are uppercase letters + digits, 6–14 chars, optional suffix.
    """
    if not isinstance(v, str):
        return False
    return bool(re.match(r'^[A-Z]{2,6}\d{4}', v.strip()))


def is_student_row(row) -> bool:
    """
    True if col 0 is a positive integer (student sequence number).
    """
    v = row[0]
    return isinstance(v, int) and v > 0


def is_course_row(row) -> bool:
    """
    True if col 0 is a non-empty string that looks like a degree name.
    Degree names contain 'bachelor' or 'doctor' or 'master' or start with 'BS'.
    """
    v = row[0]
    if not isinstance(v, str) or len(v.strip()) < 6:
        return False
    low = v.strip().lower()
    return any(kw in low for kw in ("bachelor", "doctor", "master", "diploma"))


def extract_header_info(rows) -> tuple[str, str]:
    """
    Scan the first 20 rows of a sheet for semester and academic year text.
    Returns (semester_str, academic_year_str).
    """
    semester = "1sem"
    academic_year = "Unknown"

    for row in rows[:20]:
        for cell in row:
            if cell is None:
                continue
            text = str(cell)
            for pattern, sem_val in SEM_PATTERNS.items():
                if pattern.search(text):
                    semester = sem_val
            m = YEAR_PATTERN.search(text)
            if m:
                academic_year = f"{m.group(1)}-{m.group(2)}"

    return semester, academic_year


# ══════════════════════════════════════════════════════════════════════════════
#  SHEET PARSER
# ══════════════════════════════════════════════════════════════════════════════

def parse_sheet(ws, college_name: str, semester: str, academic_year: str) -> list[dict]:
    """
    Parse one sheet and return a list of long-form dicts, one per student×subject.

    Sheet layout (repeating for each course block):
        Row N:   Course name in col 0
        Row N+2: "STUDENT NAME" header
        Row N+4: Subject codes starting at col 3
        Row N+5: Student data row (col 0 = seq#, col 1 = gender, col 3+ = grades)
        (subject row + student row pairs repeat for every student)
    """
    records = []
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return records

    # ── Find all course-block start positions ──────────────────────────────
    course_starts: list[tuple[int, str]] = []  # (row_index, course_name)
    for i, row in enumerate(rows):
        if is_course_row(row):
            course_starts.append((i, str(row[0]).strip()))

    if not course_starts:
        log.warning(f"    No course blocks found in sheet {college_name}")
        return records

    # Add sentinel so each block knows where the next starts
    course_starts.append((len(rows), "__END__"))

    # ── Process each course block ──────────────────────────────────────────
    for block_idx in range(len(course_starts) - 1):
        course_row_idx, course_name = course_starts[block_idx]
        next_course_row_idx = course_starts[block_idx + 1][0]

        block_rows = rows[course_row_idx:next_course_row_idx]

        # Within block: subject row (cols 3+) immediately precedes student row
        # We scan pairs: whenever col 0 is int, the row above it is subjects
        prev_subjects: list[str] = []

        for local_i, row in enumerate(block_rows):
            # Check if this is a subject-code row (col 3 onward)
            if row[0] is None and any(is_subject_code(v) for v in row[3:]):
                prev_subjects = [
                    str(v).strip()
                    for v in row[3:]
                    if is_subject_code(v)
                ]
                continue

            if is_student_row(row):
                student_seq = int(row[0])
                gender_raw  = str(row[1]).strip() if row[1] is not None else "Unknown"
                gender_val  = 1 if gender_raw.lower() == "female" else (
                              0 if gender_raw.lower() == "male" else -1)

                grades_raw  = row[3:]   # Col C onward

                # Pair each subject code with its grade
                for subj_i, subject_code in enumerate(prev_subjects):
                    raw_grade = grades_raw[subj_i] if subj_i < len(grades_raw) else None
                    grade_val = parse_grade(raw_grade)

                    if grade_val is None:
                        # Skip cells that are clearly not grades (None, empty, summary)
                        continue

                    # Unique stable student ID: seq#_college_course_sem_year
                    #
                    # IMPORTANT: course_tag must NOT be truncated. Every program
                    # here starts with "Bachelor of Science in ..." /
                    # "Bachelor of Arts in ..." / "Bachelor of Technical-...",
                    # so course_name[:10] always produced the same string
                    # ("Bacheloro") for every single course. That collapsed
                    # different students who happened to share a sequence
                    # number within the same college (e.g. CEA seq #1 in
                    # "Architecture" and CEA seq #1 in "Civil Engineering")
                    # into one Student_ID, silently merging two different
                    # people's grades together in every downstream groupby.
                    # Full sanitized course name = guaranteed unique per course.
                    college_tag = re.sub(r'[^A-Za-z0-9]+', '_', college_name.strip()).strip('_')
                    course_tag  = re.sub(r'[^A-Za-z0-9]+', '_', course_name.strip()).strip('_')
                    student_id  = f"{student_seq}_{college_tag}_{course_tag}_{semester}_{academic_year}"

                    records.append({
                        "Student_ID":     student_id,
                        "Student_Seq":    student_seq,
                        "Gender":         gender_val,
                        "College":        college_name,
                        "Course":         course_name,
                        "Subject":        subject_code,
                        "Grade":          grade_val,
                        "Semester":       semester,
                        "Year":           academic_year,
                    })

                # Reset subjects after consuming (each student has their own subject row)
                prev_subjects = []

    return records


# ══════════════════════════════════════════════════════════════════════════════
#  FILE PARSER
# ══════════════════════════════════════════════════════════════════════════════

def parse_workbook(filepath: str) -> pd.DataFrame:
    """Parse all sheets of one xlsx file and return long-form DataFrame."""
    log.info(f"  Parsing: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    all_records = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_preview = list(ws.iter_rows(max_row=20, values_only=True))

        # Map sheet code to full college name
        college_name = SHEET_COLLEGE_MAP.get(
            sheet_name.strip().upper(),
            sheet_name.strip()   # Fallback: use sheet name as-is
        )

        # Extract semester + year from header rows
        semester, academic_year = extract_header_info(rows_preview)

        log.info(f"    Sheet: {sheet_name} → {college_name} | {semester} | {academic_year}")

        records = parse_sheet(ws, college_name, semester, academic_year)
        log.info(f"      → {len(records):,} subject-grade records")
        all_records.extend(records)

    wb.close()
    return pd.DataFrame(all_records)


# ══════════════════════════════════════════════════════════════════════════════
#  FEATURE ENGINEERING
# ══════════════════════════════════════════════════════════════════════════════

def engineer_features(df: pd.DataFrame) -> pd.DataFrame:
    """
    From the long-form grade table, compute all per-student aggregates
    used by the ML models.
    """
    # Year numeric (start year of AY, e.g. 2022 from "2022-2023")
    df["Year_Numeric"] = (
        df["Year"].str.extract(r"^(\d{4})")[0].astype(float)
    )

    sem_map = {"1sem": 1, "2sem": 2, "summer": 3}
    df["Sem_Numeric"] = df["Semester"].map(sem_map).fillna(1)

    # Per-student aggregates (groupby student × semester × year)
    key = ["Student_ID", "Student_Seq", "Gender", "College", "Course",
           "Semester", "Year", "Year_Numeric", "Sem_Numeric"]

    student_agg = (
        df.groupby(key)["Grade"]
        .agg(
            GWA       = lambda g: g[g > 0].mean() if (g > 0).any() else np.nan,
            Avg_Grade = "mean",
            Std_Grade = "std",
            Sub_Count = "count",
            Min_Grade = "min",
            Max_Grade = "max",
        )
        .reset_index()
    )

    # Flags (per student: did they have any of these in this semester?)
    flag_df = df.groupby("Student_ID").agg(
        is_inc       = ("Grade", lambda g: int((g == 5.0).any())),
        is_drop      = ("Grade", lambda g: int((g == 0.0).any())),
        fail_count   = ("Grade", lambda g: int((g >= 3.0).sum())),
    ).reset_index()

    student_agg = student_agg.merge(flag_df, on="Student_ID", how="left")

    # Derived rate columns
    student_agg["fail_rate"]     = student_agg["fail_count"] / student_agg["Sub_Count"].replace(0, np.nan)
    student_agg["inc_rate"]      = student_agg["is_inc"]  # binary per student
    student_agg["drop_rate"]     = student_agg["is_drop"]
    student_agg["is_irregular"]  = (
        (student_agg["is_inc"] == 1) | (student_agg["is_drop"] == 1)
    ).astype(int)

    return student_agg, df   # return both aggregated and raw long-form


# ══════════════════════════════════════════════════════════════════════════════
#  MODEL DATASET BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def build_model_datasets(student_df: pd.DataFrame, long_df: pd.DataFrame, out_dir: str):
    """Build all 12 model-dataset CSVs."""
    os.makedirs(out_dir, exist_ok=True)

    def save(df, name):
        path = os.path.join(out_dir, name)
        df.to_csv(path, index=False)
        log.info(f"    Saved {name}: {len(df):,} rows")

    # 01 – Dropout risk per student (the main student-level table)
    save(student_df, "01_dropout_risk_per_student.csv")

    # 02 – Dropout spike cohort (college × year dropout rate)
    spike = (
        student_df.groupby(["Year_Numeric", "College"])
        .agg(
            Total_Students = ("Student_ID", "nunique"),
            Dropout_Count  = ("is_drop", "sum"),
        )
        .reset_index()
    )
    spike["Dropout_Rate"]    = (spike["Dropout_Count"] / spike["Total_Students"] * 100).round(2)
    spike["Non_Dropout_Pct"] = (100 - spike["Dropout_Rate"]).round(2)
    save(spike, "02_dropout_spike_cohort.csv")

    # 03 – Dropout ranking college (student-level with key columns)
    save(
        student_df[["Student_ID", "College", "Course", "Semester",
                    "Sem_Numeric", "Year_Numeric", "GWA", "fail_rate", "is_drop"]],
        "03_dropout_ranking_college.csv"
    )

    # 04 – GWA ranking college (student-level)
    save(
        student_df[["Student_ID", "College", "Course",
                    "Year_Numeric", "Sem_Numeric", "GWA"]].dropna(subset=["GWA"]),
        "04_gwa_ranking_college.csv"
    )

    # 05 – GWA trend timeseries (college × year × sem)
    trend = (
        student_df.dropna(subset=["GWA"])
        .groupby(["Year_Numeric", "Sem_Numeric", "College"])
        .agg(
            Avg_GWA     = ("GWA", "mean"),
            Std_GWA     = ("GWA", "std"),
            Student_Cnt = ("Student_ID", "nunique"),
        )
        .reset_index()
    )
    trend["Avg_GWA"] = trend["Avg_GWA"].round(2)
    trend["Std_GWA"] = trend["Std_GWA"].round(2)
    save(trend, "05_gwa_trend_timeseries.csv")

    # 06 – INC forecast cohort
    inc = (
        student_df.groupby(["Year_Numeric", "Sem_Numeric", "College"])
        .agg(
            Total_Students = ("Student_ID", "nunique"),
            INC_Count      = ("is_inc", "sum"),
        )
        .reset_index()
    )
    inc["INC_Rate"] = (inc["INC_Count"] / inc["Total_Students"] * 100).round(2)
    save(inc, "06_inc_forecast_cohort.csv")

    # 07 – Irreg/Reg cohort
    irreg = (
        student_df.groupby(["Year_Numeric", "Sem_Numeric", "College"])
        .agg(
            Total_Students  = ("Student_ID", "nunique"),
            Irregular_Count = ("is_irregular", "sum"),
            Drop_Count      = ("is_drop", "sum"),
            INC_Count       = ("is_inc", "sum"),
        )
        .reset_index()
    )
    irreg["Irregular_Rate"] = (irreg["Irregular_Count"] / irreg["Total_Students"] * 100).round(2)
    irreg["Drop_Rate"]      = (irreg["Drop_Count"]      / irreg["Total_Students"] * 100).round(2)
    irreg["INC_Rate"]       = (irreg["INC_Count"]       / irreg["Total_Students"] * 100).round(2)
    save(irreg, "07_irreg_reg_cohort.csv")

    # 08 – KPI GWA student (same as 04)
    save(
        student_df[["Student_ID", "College", "Course",
                    "Year_Numeric", "Sem_Numeric", "GWA"]].dropna(subset=["GWA"]),
        "08_kpi_gwa_student.csv"
    )

    # 09 – KPI enrollment college
    enroll = (
        student_df.groupby(["Year_Numeric", "Sem_Numeric", "College"])
        .agg(Headcount=("Student_ID", "nunique"))
        .reset_index()
        .sort_values(["College", "Year_Numeric", "Sem_Numeric"])
    )
    enroll["Headcount_Prev"] = enroll.groupby("College")["Headcount"].shift(1)
    enroll["Growth_Rate"] = (
        (enroll["Headcount"] - enroll["Headcount_Prev"])
        / enroll["Headcount_Prev"] * 100
    ).round(2)
    save(enroll, "09_kpi_enrollment_college.csv")

    # 10 – Subject grade forecast (long_df aggregated per subject)
    subj = (
        long_df[long_df["Grade"] > 0]   # exclude drops/INCs for grade difficulty
        .groupby(["Year_Numeric", "College", "Course", "Subject"])
        .agg(
            Avg_Grade   = ("Grade", "mean"),
            Std_Grade   = ("Grade", "std"),
            Student_Cnt = ("Student_ID", "nunique"),
            Fail_Count  = ("Grade", lambda g: (g >= 3.0).sum()),
        )
        .reset_index()
    )
    subj["Avg_Grade"] = subj["Avg_Grade"].round(2)
    subj["Std_Grade"] = subj["Std_Grade"].round(2)
    subj["Fail_Rate"] = (subj["Fail_Count"] / subj["Student_Cnt"] * 100).round(2)
    save(subj, "10_subject_grade_forecast.csv")

    # 11 – Performance band distribution
    def perf_band(gwa):
        if pd.isna(gwa): return "Unknown"
        if gwa <= 1.5:   return "Excellent"
        if gwa <= 2.0:   return "Good"
        if gwa <= 2.5:   return "Average"
        if gwa <= 3.0:   return "Below Average"
        return "Failing"

    band_df = student_df.dropna(subset=["GWA"]).copy()
    band_df["Perf_Band"] = band_df["GWA"].apply(perf_band)
    band = (
        band_df.groupby(["Year_Numeric", "Sem_Numeric", "College", "Perf_Band"])
        .agg(Count=("Student_ID", "nunique"))
        .reset_index()
    )
    total_map = (
        band_df.groupby(["Year_Numeric", "Sem_Numeric", "College"])
        ["Student_ID"].nunique()
        .rename("Total")
        .reset_index()
    )
    band = band.merge(total_map, on=["Year_Numeric", "Sem_Numeric", "College"], how="left")
    band["Pct"] = (band["Count"] / band["Total"] * 100).round(2)
    save(band, "11_performance_band_dist.csv")

    # 12 – Gender performance
    gender = (
        student_df.groupby(["Year_Numeric", "College", "Gender"])
        .agg(
            Student_Count  = ("Student_ID", "nunique"),
            Avg_GWA        = ("GWA", "mean"),
            Dropout_Rate   = ("is_drop", lambda g: g.mean() * 100),
            INC_Rate       = ("is_inc",  lambda g: g.mean() * 100),
            Irregular_Rate = ("is_irregular", lambda g: g.mean() * 100),
        )
        .reset_index()
    )
    gender["Avg_GWA"]        = gender["Avg_GWA"].round(2)
    gender["Dropout_Rate"]   = gender["Dropout_Rate"].round(2)
    gender["INC_Rate"]       = gender["INC_Rate"].round(2)
    gender["Irregular_Rate"] = gender["Irregular_Rate"].round(2)
    gender["Gender_Label"]   = gender["Gender"].map({1: "Female", 0: "Male", -1: "Unknown"})
    save(gender, "12_gender_performance.csv")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def run_preprocessing(
    input_dir:  str,
    output_dir: str,
    model_datasets_dir: str | None = None,
) -> dict:
    """
    Main preprocessing pipeline.  Called by upload_routes.py after a file is
    saved to disk, OR run standalone via CLI.

    Parameters
    ----------
    input_dir           Folder containing all uploaded .xlsx files
    output_dir          Where to write Final_Merged_Student_Data.csv + long-form CSV
    model_datasets_dir  Where to write the 12 model-dataset CSVs
                        (defaults to output_dir/model_datasets)

    Returns
    -------
    dict with keys: success, student_rows, subject_rows, files_processed, error
    """
    if model_datasets_dir is None:
        model_datasets_dir = os.path.join(output_dir, "model_datasets")

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(model_datasets_dir, exist_ok=True)

    xlsx_files = sorted(Path(input_dir).glob("*.xlsx"))
    if not xlsx_files:
        return {"success": False, "error": "No .xlsx files found in input directory"}

    log.info(f"Found {len(xlsx_files)} xlsx files in {input_dir}")

    all_long_dfs = []

    for fpath in xlsx_files:
        try:
            df = parse_workbook(str(fpath))
            if not df.empty:
                all_long_dfs.append(df)
        except Exception as e:
            log.error(f"  Failed to parse {fpath.name}: {e}")
            continue

    if not all_long_dfs:
        return {"success": False, "error": "All files failed to parse"}

    # Concatenate ALL long-form records
    long_df = pd.concat(all_long_dfs, ignore_index=True)
    log.info(f"Total long-form records: {len(long_df):,}")
    log.info(f"Unique students:         {long_df['Student_ID'].nunique():,}")
    log.info(f"Unique subjects:         {long_df['Subject'].nunique():,}")

    # Save the raw long-form (subject-level) master CSV
    long_csv_path = os.path.join(output_dir, "Final_LongForm_Student_Grades.csv")
    long_df.to_csv(long_csv_path, index=False)
    log.info(f"Saved long-form CSV: {long_csv_path}")

    # Feature engineering → student-level aggregates
    student_df, long_df = engineer_features(long_df)
    log.info(f"Student-level rows: {len(student_df):,}")

    # Save the student-level master CSV (used by ml_analysis.py as FINAL_MERGED_CSV)
    merged_csv_path = os.path.join(output_dir, "Final_Merged_Student_Data.csv")
    student_df.to_csv(merged_csv_path, index=False)
    log.info(f"Saved student CSV: {merged_csv_path}")

    # Build and save all 12 model datasets
    log.info("Building model datasets...")
    build_model_datasets(student_df, long_df, model_datasets_dir)

    return {
        "success":         True,
        "student_rows":    len(student_df),
        "subject_rows":    len(long_df),
        "files_processed": len(xlsx_files),
        "error":           None,
    }


# ══════════════════════════════════════════════════════════════════════════════
#  COMPATIBILITY SHIM FOR auto_train.py
#  ─────────────────────────────────────────────────────────────────────────────
#  auto_train.py imports these five names from this module:
#
#    from preprocessing.preprocess import (
#        process_file, export_model_datasets,
#        FINAL_COLUMNS, PROCESSED_DIR, MODEL_DATA_DIR, FINAL_OUTPUT,
#    )
#
#  They map to the new API as follows:
# ══════════════════════════════════════════════════════════════════════════════

try:
    from configs.config import (
        PROCESSED_DATASETS_DIR  as _PROCESSED_DIR,
        MODEL_DATASETS_DIR      as _MODEL_DATA_DIR,
        FINAL_MERGED_CSV        as _FINAL_OUTPUT,
        UNPROCESSED_DATASETS_DIR as _UNPROCESSED_DIR,
    )
except ImportError:
    # Fallback for standalone / CLI use outside the Flask app
    _PROCESSED_DIR   = os.path.join(os.path.dirname(__file__), "..", "Processed_Datasets")
    _MODEL_DATA_DIR  = os.path.join(_PROCESSED_DIR, "model_datasets")
    _FINAL_OUTPUT    = os.path.join(_PROCESSED_DIR, "Final_Merged_Student_Data.csv")
    _UNPROCESSED_DIR = os.path.join(os.path.dirname(__file__), "..", "Unprocessed_Datasets")

# ── Constants ─────────────────────────────────────────────────────────────────
PROCESSED_DIR   = _PROCESSED_DIR
MODEL_DATA_DIR  = _MODEL_DATA_DIR
FINAL_OUTPUT    = _FINAL_OUTPUT

# Columns that auto_train.py keeps when it de-dupes the master CSV.
# Must match the student-level CSV produced by engineer_features().
FINAL_COLUMNS = [
    "Student_ID", "Student_Seq", "Gender", "College", "Course",
    "Semester", "Year", "Year_Numeric", "Sem_Numeric",
    "GWA", "Avg_Grade", "Std_Grade", "Sub_Count",
    "Min_Grade", "Max_Grade",
    "is_inc", "is_drop", "fail_count", "fail_rate",
    "inc_rate", "drop_rate", "is_irregular",
]


def process_file(xlsx_path: str) -> pd.DataFrame:
    """
    Compatibility wrapper called by auto_train.run_full_pipeline(new_file=...).

    Parses a single .xlsx file and returns a student-level DataFrame with
    all FINAL_COLUMNS populated.  The caller (auto_train) then merges this
    with the existing master CSV, de-dupes, and calls export_model_datasets().

    Also persists the long-form (subject-level) records to
    Final_LongForm_Student_Grades.csv, merging with whatever is already
    there. This file previously was only ever written by the standalone
    run_preprocessing() CLI path — which the upload pipeline never calls —
    so on a fresh install export_model_datasets() always fell back to an
    empty long_df, dataset 10 (subject_grade_forecast) always came out with
    0 rows, and train_subject_top() always skipped with "too few aggregated
    rows" no matter how much real data had been uploaded.
    """
    log.info(f"process_file: {xlsx_path}")

    long_df = parse_workbook(xlsx_path)
    if long_df.empty:
        log.warning("process_file: no records extracted — returning empty DataFrame")
        return pd.DataFrame(columns=FINAL_COLUMNS)

    # ── Persist long-form rows so export_model_datasets() can read them ────
    # Add Year_Numeric before saving: export_model_datasets() re-reads this
    # file fresh from disk on a later, possibly separate call, so it can't
    # rely on the in-place mutation engineer_features() does below to the
    # in-memory long_df — the saved CSV needs the column itself.
    long_df["Year_Numeric"] = (
        long_df["Year"].astype(str).str.extract(r"^(\d{4})")[0].astype(float)
    )

    long_csv_path = os.path.join(PROCESSED_DIR, "Final_LongForm_Student_Grades.csv")
    try:
        os.makedirs(PROCESSED_DIR, exist_ok=True)
        if os.path.exists(long_csv_path):
            existing_long = pd.read_csv(long_csv_path)
            combined_long = pd.concat([existing_long, long_df], ignore_index=True)
        else:
            combined_long = long_df
        # Same de-dupe key auto_train.py uses for the student-level master CSV
        dedupe_cols = [c for c in ["Student_ID", "Subject", "Semester", "Year"]
                       if c in combined_long.columns]
        if dedupe_cols:
            combined_long = combined_long.drop_duplicates(subset=dedupe_cols, keep="last")
        combined_long.to_csv(long_csv_path, index=False)
        log.info(f"  Long-form CSV updated: {len(combined_long):,} rows -> {long_csv_path}")
    except Exception as e:
        log.error(f"  Failed to persist long-form CSV: {e}")

    student_df, _ = engineer_features(long_df)

    # Ensure all FINAL_COLUMNS exist (fill missing with sensible defaults)
    for col in FINAL_COLUMNS:
        if col not in student_df.columns:
            student_df[col] = np.nan

    return student_df[FINAL_COLUMNS]


def export_model_datasets(merged_df: pd.DataFrame, out_dir: str):
    """
    Compatibility wrapper called by auto_train after it updates the master CSV.

    Rebuilds the long-form subject table from the master student CSV and
    writes all 12 model-dataset CSVs to out_dir.

    Note: because auto_train passes a student-level (aggregated) DataFrame,
    we use it directly for the student-based datasets (01–09, 11–12) and
    re-read the long-form CSV for subject-level data (10) if it exists.
    """
    log.info(f"export_model_datasets → {out_dir}")

    # Try to load the long-form CSV for subject-level dataset (10)
    long_csv = os.path.join(PROCESSED_DIR, "Final_LongForm_Student_Grades.csv")
    if os.path.exists(long_csv):
        long_df = pd.read_csv(long_csv)
    else:
        # Fallback: create a minimal long_df from student-level data
        # (subject-level detail will be empty but won't crash)
        log.warning("Long-form CSV not found; subject dataset (10) will be empty")
        long_df = pd.DataFrame(columns=["Student_ID", "College", "Course",
                                         "Subject", "Grade", "Year_Numeric"])

    build_model_datasets(merged_df, long_df, out_dir)


# ── CLI ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="NovaSight Grade Preprocessor v2")
    parser.add_argument("--input_dir",  default="Unprocessed_Datasets",
                        help="Folder with .xlsx grade sheets")
    parser.add_argument("--output_dir", default="Processed_Datasets",
                        help="Output folder for CSVs")
    parser.add_argument("--model_dir",  default=None,
                        help="Output folder for model datasets (default: output_dir/model_datasets)")
    args = parser.parse_args()

    result = run_preprocessing(args.input_dir, args.output_dir, args.model_dir)

    if result["success"]:
        print("\nPreprocessing complete!")
        print(f"   Files processed : {result['files_processed']}")
        print(f"   Students        : {result['student_rows']:,}")
        print(f"   Subject records : {result['subject_rows']:,}")
    else:
        print(f"\nPreprocessing failed: {result['error']}")
        sys.exit(1)