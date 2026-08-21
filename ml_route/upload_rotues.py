import os
import re
import time
import glob
import threading
import shutil

from flask import Blueprint, request, jsonify, session, render_template, current_app
from werkzeug.utils import secure_filename

from database.models import db, AcadUser, UploadedDataset
from configs.config import (
    UNPROCESSED_DATASETS_DIR,
    PROCESSED_DATASETS_DIR,
    MODEL_DATASETS_DIR,
    FINAL_MERGED_CSV,
    DATASET_ALLOWED_EXTENSIONS,
    DATASET_MAX_SIZE_MB,
    DATASET_FILENAME_REGEX,
    UPLOAD_ALLOWED_ROLES,
)
from training.auto_train import run_full_pipeline, load_state, MODEL_DIR as ML_MODEL_DIR

upload_bp = Blueprint('upload_bp', __name__)


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def _current_user():
    uid = session.get('user_id')
    return AcadUser.query.get(uid) if uid else None


def _validate_dataset_file(filename: str) -> tuple[bool, str]:
    """
    Two-rule validation:
      1. Extension must be .xlsx
      2. Filename must match DATASET_FILENAME_REGEX
         (YYYY-N[_ ]Student-Performance[_ ]Dataset.xlsx)
    No file size limit.
    """
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in DATASET_ALLOWED_EXTENSIONS:
        return False, (
            f"Invalid file type <code>.{ext}</code>. "
            "Only <strong>.xlsx</strong> grade-sheet workbooks are accepted."
        )

    if not DATASET_FILENAME_REGEX.match(filename):
        return False, (
            "Filename does not match the required format.<br>"
            "Expected: <strong>YYYY-N Student-Performance Dataset.xlsx</strong><br>"
            "Where <strong>N</strong> is <code>1</code> (1st sem) or <code>2</code> (2nd sem).<br>"
            "Examples:<br>"
            "&nbsp;&nbsp;• <code>2022-1 Student-Performance Dataset.xlsx</code><br>"
            "&nbsp;&nbsp;• <code>2025-2 Student-Performance Dataset.xlsx</code><br>"
            f"Received: <code>{filename}</code>"
        )

    return True, ''


def _parse_filename_meta(filename: str) -> tuple[str, str]:
    """
    Parse academic year and semester from filename.
    '2022-1 Student-Performance Dataset.xlsx'  →  ('2022-2023', '1sem')
    '2023-2_Student-Performance_Dataset.xlsx'  →  ('2023-2024', '2sem')
    """
    m = re.match(r'(\d{4})-([12])', filename)
    if m:
        yr  = int(m.group(1))
        sem = int(m.group(2))
        return f"{yr}-{yr + 1}", f"{sem}sem"
    return 'Unknown', '1sem'


def _safe_stored_name(user_id: int, original: str) -> str:
    """Collision-safe stored filename: userid_timestamp_safename."""
    # Normalise spaces → underscores before securing
    normalized = original.replace(' ', '_')
    return f"{user_id}_{int(time.time())}_{secure_filename(normalized)}"


def _canonical_name(filename: str) -> str:
    """
    Return a normalised canonical name used for duplicate detection.
    Spaces and underscores are treated as equivalent.
    e.g. '2022-1 Student-Performance Dataset.xlsx'
      →  '2022-1_Student-Performance_Dataset.xlsx'
    """
    return filename.replace(' ', '_')


# ─────────────────────────────────────────────────────────────
# BACKGROUND WORKER
# ─────────────────────────────────────────────────────────────

def _background_train(app, record_id: int, raw_path: str):
    """
    Background thread:
      1. Run full preprocessing + model retraining pipeline
      2. Update DB record on completion
    """
    with app.app_context():
        record = UploadedDataset.query.get(record_id)
        if not record:
            return

        record.status = 'processing'
        db.session.commit()

        try:
            state = run_full_pipeline(new_file=raw_path)

            row_count  = state.get('rows_in_master')
            label      = (record.academic_year or '').replace('-', '_')
            sem_csv    = os.path.join(
                PROCESSED_DATASETS_DIR,
                f"{label}_{record.semester}_cleaned.csv"
            )
            proc_path  = sem_csv if os.path.exists(sem_csv) else FINAL_MERGED_CSV

            record.processed      = True
            record.status         = 'done'
            record.processed_path = proc_path
            record.row_count      = row_count
            db.session.commit()

            # Hot-reload ML models so the dashboard reflects the new PKLs
            # immediately without requiring a Flask restart.
            try:
                from ml_route.ml_analysis import reload_models
                reload_models()
            except Exception as _re:
                print(f"[upload_routes] reload_models warning: {_re}")

        except Exception as exc:
            import traceback
            record.status        = 'failed'
            record.error_message = str(exc)
            db.session.commit()
            traceback.print_exc()

            # ── Clean up the physical files ─────────────────────
            # The duplicate check only looks at whether a canonical-name
            # copy exists on disk. If we left that copy sitting in
            # Unprocessed_Datasets/ after a failed run, re-uploading the
            # exact same filename would be rejected as a "duplicate"
            # forever, even though nothing ever succeeded. Remove both
            # the raw upload and its canonical copy so a reupload of the
            # same filename is accepted right away. The DB record itself
            # is kept (status='failed') only long enough to drive the
            # "file failed to process" floating notice on the frontend —
            # see /api/upload-record/<id> for how it's removed for good.
            canonical_path = os.path.join(
                UNPROCESSED_DATASETS_DIR, _canonical_name(record.original_filename)
            )
            for stale_path in (raw_path, canonical_path):
                if stale_path and os.path.exists(stale_path):
                    try:
                        os.remove(stale_path)
                    except OSError as cleanup_err:
                        print(f"[upload_routes] failed-file cleanup warning: {cleanup_err}")


# ─────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────

@upload_bp.route('/upload-file')
def upload_file_page():
    user = _current_user()
    return render_template('studentaffair/fileupload/fileupload.html', user=user)


@upload_bp.route('/api/upload-dataset', methods=['POST'])
def api_upload_dataset():
    user = _current_user()
    if not user:
        return jsonify({'ok': False, 'error': 'Not logged in.'}), 401
    if user.role not in UPLOAD_ALLOWED_ROLES:
        return jsonify({'ok': False, 'error': 'Your role is not permitted to upload datasets.'}), 403

    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file included in the request.'}), 400

    f = request.files['file']
    if not f.filename:
        return jsonify({'ok': False, 'error': 'Empty filename.'}), 400

    # Measure size for display only (no limit enforced)
    f.seek(0, 2)
    size_bytes = f.tell()
    f.seek(0)

    # ── Format validation (extension + filename pattern only) ──
    ok, reason = _validate_dataset_file(f.filename)
    if not ok:
        return jsonify({'ok': False, 'error': reason}), 422

    # ── Duplicate check (canonical name in Unprocessed_Datasets/) ──
    canonical = _canonical_name(f.filename)
    orig_path = os.path.join(UNPROCESSED_DATASETS_DIR, canonical)

    if os.path.exists(orig_path):
        dup = UploadedDataset.query.filter(
            UploadedDataset.original_filename.in_([f.filename, canonical])
        ).first()
        return jsonify({
            'ok'       : False,
            'duplicate': True,
            'error'    : (
                f"<strong>'{f.filename}'</strong> has already been uploaded"
                + (
                    f" by <strong>{dup.uploader.username}</strong>"
                    f" on {dup.uploaded_at.strftime('%b %d, %Y')}"
                    if dup else ''
                )
                + ".<br>If this is a different dataset please rename the file and re-upload."
            ),
        }), 409

    # ── Save raw file ─────────────────────────────────────────
    stored_name = _safe_stored_name(user.acaduser_id, f.filename)
    raw_path    = os.path.join(UNPROCESSED_DATASETS_DIR, stored_name)
    f.save(raw_path)

    # Keep a canonical-name copy for future duplicate checks
    shutil.copy2(raw_path, orig_path)

    year, semester = _parse_filename_meta(f.filename)
    size_kb = round(size_bytes / 1024, 1)

    # ── Count sheets ─────────────────────────────────────────
    sheet_count = None
    try:
        from openpyxl import load_workbook
        wb = load_workbook(raw_path, read_only=True)
        sheet_count = len(wb.sheetnames)
        wb.close()
    except Exception:
        pass

    # ── Create DB record ──────────────────────────────────────
    record = UploadedDataset(
        original_filename = f.filename,
        stored_filename   = stored_name,
        raw_path          = raw_path,
        status            = 'pending',
        uploaded_by       = user.acaduser_id,
        academic_year     = year,
        semester          = semester,
        file_size_kb      = size_kb,
        sheet_count       = sheet_count,
    )
    db.session.add(record)
    db.session.commit()

    # ── Launch background training ────────────────────────────
    app    = current_app._get_current_object()
    thread = threading.Thread(
        target=_background_train,
        args=(app, record.id, raw_path),
        daemon=True,
    )
    thread.start()

    return jsonify({
        'ok'       : True,
        'record_id': record.id,
        'message'  : (
            f"'{f.filename}' accepted. "
            "Preprocessing and model retraining have started in the background."
        ),
    }), 202


@upload_bp.route('/api/upload-record/<int:record_id>', methods=['DELETE'])
def api_delete_upload_record(record_id: int):
    """
    Permanently remove an upload record — used by the 'Remove' button on
    the failed-upload floating card. Only 'failed' or still-'pending'
    records may be removed this way; anything mid-flight ('processing')
    or already 'done' should go through the reset-all-data flow instead,
    not be silently deleted one row at a time.
    """
    user = _current_user()
    if not user:
        return jsonify({'ok': False, 'error': 'Not logged in.'}), 401
    if user.role not in UPLOAD_ALLOWED_ROLES:
        return jsonify({'ok': False, 'error': 'Your role is not permitted to modify uploads.'}), 403

    record = UploadedDataset.query.get(record_id)
    if not record:
        return jsonify({'ok': False, 'error': 'Record not found.'}), 404

    if record.status not in ('failed', 'pending'):
        return jsonify({
            'ok': False,
            'error': f"Cannot remove a record with status '{record.status}'.",
        }), 400

    # Belt-and-suspenders: the failure handler already deletes these, but
    # clear any leftover files in case this record failed before that
    # cleanup existed, or is still 'pending'.
    canonical_path = os.path.join(
        UNPROCESSED_DATASETS_DIR, _canonical_name(record.original_filename)
    )
    for stale_path in (record.raw_path, canonical_path):
        if stale_path and os.path.exists(stale_path):
            try:
                os.remove(stale_path)
            except OSError as cleanup_err:
                print(f"[upload_routes] remove-record cleanup warning: {cleanup_err}")

    db.session.delete(record)
    db.session.commit()

    return jsonify({'ok': True, 'message': 'Upload record removed.'})


@upload_bp.route('/api/failed-uploads')
def api_failed_uploads():
    """
    Any 'failed' upload records still sitting in the DB (not yet dismissed
    via the Remove button). Polled once on page load so the failed-upload
    floating card can resurface after a refresh instead of only appearing
    during the same in-flight upload session.
    """
    records = (
        UploadedDataset.query
        .filter_by(status='failed')
        .order_by(UploadedDataset.uploaded_at.desc())
        .all()
    )
    return jsonify([r.to_dict() for r in records])


@upload_bp.route('/api/upload-status/<int:record_id>')
def api_upload_status(record_id: int):
    """Poll until status is 'done' or 'failed'."""
    record = UploadedDataset.query.get_or_404(record_id)
    data   = record.to_dict()

    # Attach prediction horizon from training_state.json
    state = load_state()
    if state.get('horizon'):
        data['horizon'] = state['horizon']

    return jsonify(data)


@upload_bp.route('/api/training-state')
def api_training_state():
    """Full training_state.json — model metrics + prediction horizon."""
    state = load_state()
    if not state:
        return jsonify({'status': 'no_training_yet'}), 200
    return jsonify(state)


# ── Human-readable labels for each trained model ────────────────
# For models that train several sub-models at once (kpi, gender_performance_*),
# this is the SHARED prefix — the actual sub-model name comes from
# _SUBMODEL_LABELS below and each sub-model gets rendered as its own
# dashboard card instead of one bundled card.
#
# performance_band, year_level_performance, and year_level_inc_irreg were
# REMOVED (2026-08-19) from auto_train.py's trainer list — all three were
# RandomForestRegressor models trained to eventually power a forecast, but
# RF can't extrapolate past the years it was trained on (see forecast_series()'s
# docstring in ml_analysis.py — the same failure mode already hit and fixed
# once for the subject-grade forecast). year_level_inc_irreg was additionally
# measurably harmful (Drop_Rate R^2 = -0.63) and fully redundant with
# /api/get_year_level_inc_irreg_forecast, which already forecasts all three
# rates live via forecast_series(). Deliberately not listed here anymore —
# if a stale training_state.json still has them, they'll fall back to a
# Title-Cased key and an unwired "chart_used" badge, which is the correct
# signal to retrain and drop them for good.
_MODEL_LABELS = {
    'dropout_risk':       'Dropout Risk (per Student)',
    'dropout_spike':      'Dropout Spike (Cohort Trend)',
    'dropout_ranking':    'Dropout Ranking (per College)',
    'gwa_ranking':        'GWA Ranking (per College)',
    'gwa_trend':          'GWA Trend (Time-Series)',
    'inc_forecast':       'INC Rate Forecast',
    'irreg_reg':          'Irregular vs Regular (per Student)',
    'kpi':                'KPI',
    'subject_grade':      'Subject Grade Forecast',
    'gender_performance_male':   'Gender Performance — Male',
    'gender_performance_female': 'Gender Performance — Female',
}

# Sub-model display names, keyed by parent model key -> {sub_name: label}.
# Only needed for trainers that return a nested dict (one dict per sub-model).
_SUBMODEL_LABELS = {
    'kpi': {'gwa': 'GWA', 'enrollment': 'Enrollment', 'drop': 'Drop Rate'},
    'gender_performance_male':   {'dropout_rate': 'Dropout Rate', 'inc_rate': 'INC Rate'},
    'gender_performance_female': {'dropout_rate': 'Dropout Rate', 'inc_rate': 'INC Rate'},
}

# Which chart/endpoint actually consumes each model's predictions, keyed by
# the FULL model key (parent key, or "parent_subname" for split sub-models).
# None/absent = trained but not wired to anything yet — surfaced on the
# dashboard card instead of silently hidden, so a dead model is visible
# the moment it happens instead of needing another grep session to find.
# Short chart NAME each model's predictions feed — deliberately short
# (no endpoint paths/descriptions) since this gets folded straight into
# the card title as "<Model Label> — <Chart Name>", same pattern as the
# "KPI — GWA" sub-model labels, rather than shown as a separate row.
# None/absent = trained but not wired to any chart yet.
_CHART_USAGE = {
    'dropout_risk':    'Student Status Donut',
    'dropout_spike':   'Dropout Trend & Spike Chart',
    'dropout_ranking': 'College Dropout Ranking',
    'gwa_ranking':     'GWA Ranking Bar Chart',
    'gwa_trend':       'GWA Trend Line Chart',
    'inc_forecast':    'INC Rate Forecast Chart',
    'irreg_reg':       'Irregular-Rate Donut (Forecast Mode)',
    'subject_grade':   'Subject Forecast / Hardest Subjects Chart',
    'kpi_gwa':         'KPI Tile',
    'kpi_enrollment':  'KPI Tile',
    'kpi_drop':        'KPI Tile',
    'gender_performance_male_dropout_rate':   'Retention & Risk Donut (Male, per-college)',
    'gender_performance_male_inc_rate':       'Retention & Risk Donut (Male, per-college)',
    'gender_performance_female_dropout_rate': 'Retention & Risk Donut (Female, per-college)',
    'gender_performance_female_inc_rate':     'Retention & Risk Donut (Female, per-college)',
}


def _titled_label(key: str, label: str) -> str:
    """Fold the chart-usage name straight into the card title: '<label> — <chart>',
    or '<label> — Not Used in Any Chart Yet' if nothing consumes this model —
    so an unwired model is impossible to miss without a separate badge row."""
    chart = _CHART_USAGE.get(key)
    return f"{label} — {chart}" if chart else f"{label} — Not Used in Any Chart Yet"


def _build_flat_entry(key: str, label: str, result: dict) -> dict:
    """Build one dashboard-card entry for a flat (non-nested) result dict."""

    status = result.get('status', 'ok' if 'error' not in result else 'error')
    metrics = {k: v for k, v in result.items() if k not in ('status', 'reason', 'error')}

    if 'accuracy' in metrics:
        headline_label, headline_value = 'Accuracy', metrics['accuracy']
    elif 'r2' in metrics:
        headline_label, headline_value = 'R² Score', metrics['r2']
    else:
        headline_label, headline_value = None, None

    return {
        'key': key,
        'label': _titled_label(key, label),
        'status': status,
        'headline_label': headline_label,
        'headline_value': headline_value,
        'metrics': metrics,
        'reason': result.get('reason') or result.get('error'),
    }


def _flatten_metric_block(key: str, result: dict) -> list:
    """
    Normalise a trainer's result dict into one or more dashboard-card
    entries: status, a primary headline metric (accuracy if classification,
    R^2 if regression), and the rest as secondary metrics.

    Some trainers return a NESTED dict — one sub-result per sub-model
    (train_kpi's 'gwa'/'enrollment'/'drop', train_gender_performance's
    'dropout_rate'/'inc_rate'). Each sub-model gets its OWN entry/card here
    (rather than one bundled card with prefixed metric keys) so each
    sub-model's accuracy is visible on its own, instead of averaging
    distinct sub-models' evals into one card. Which chart each card feeds
    is baked directly into its title via _titled_label() (e.g.
    "KPI — GWA — KPI Tile") rather than shown as a separate row.
    """
    if not isinstance(result, dict):
        return [{'key': key, 'label': _titled_label(key, key), 'status': 'unknown',
                  'headline_label': None, 'headline_value': None, 'metrics': {}}]

    has_nested_submodels = any(isinstance(v, dict) for v in result.values())
    if not has_nested_submodels:
        label = _MODEL_LABELS.get(key, key.replace('_', ' ').title())
        return [_build_flat_entry(key, label, result)]

    base_label = _MODEL_LABELS.get(key, key.replace('_', ' ').title())
    sub_labels = _SUBMODEL_LABELS.get(key, {})
    entries = []
    for sub_name, sub_result in result.items():
        if not isinstance(sub_result, dict):
            continue
        sub_key = f"{key}_{sub_name}"
        sub_label = sub_labels.get(sub_name, sub_name.replace('_', ' ').title())
        entries.append(_build_flat_entry(sub_key, f"{base_label} — {sub_label}", sub_result))
    return entries or [_build_flat_entry(key, base_label, {'status': 'skipped'})]


@upload_bp.route('/api/model-performance')
def api_model_performance():
    """
    Dashboard-ready model evaluation summary.
    Reshapes training_state.json into a flat list, one entry per model,
    each with a headline accuracy/R² figure plus the full metric set —
    so the frontend can render a 'Model Performance' card without having
    to know each trainer's individual result shape.
    """
    state = load_state()
    if not state:
        return jsonify({
            'status': 'no_training_yet',
            'message': 'No model has been trained yet. Upload a dataset to begin.',
            'models': [],
        }), 200

    models = []
    for key, result in state.get('models', {}).items():
        models.extend(_flatten_metric_block(key, result))

    return jsonify({
        'status': 'ok',
        'trained_at': state.get('trained_at'),
        'triggered_by': state.get('triggered_by'),
        'rows_in_master': state.get('rows_in_master'),
        'elapsed_seconds': state.get('elapsed_seconds'),
        'errors': state.get('errors', []),
        'horizon': state.get('horizon', {}),
        'models': models,
    })


@upload_bp.route('/api/unprocessed-list')
def api_unprocessed_list():
    records = (
        UploadedDataset.query
        .filter(UploadedDataset.status != 'failed')
        .order_by(UploadedDataset.uploaded_at.desc())
        .all()
    )
    return jsonify([r.to_dict() for r in records])


@upload_bp.route('/api/processed-list')
def api_processed_list():
    records = (
        UploadedDataset.query
        .filter_by(status='done')
        .order_by(UploadedDataset.uploaded_at.desc())
        .all()
    )

    # Model dataset files on disk
    model_files = []
    if os.path.isdir(MODEL_DATASETS_DIR):
        for fname in sorted(os.listdir(MODEL_DATASETS_DIR)):
            if fname.endswith('.csv'):
                fpath = os.path.join(MODEL_DATASETS_DIR, fname)
                model_files.append({
                    'filename': fname,
                    'size_kb' : f"{os.path.getsize(fpath) / 1024:.1f} KB",
                    'modified': time.strftime(
                        '%b %d, %Y  %I:%M %p',
                        time.localtime(os.path.getmtime(fpath))
                    ),
                })

    state   = load_state()
    horizon = state.get('horizon', {})

    return jsonify({
        'uploaded_records'  : [r.to_dict() for r in records],
        'model_files'       : model_files,
        'merged_csv_exists' : os.path.exists(FINAL_MERGED_CSV),
        'horizon'           : horizon,
    })


# ─────────────────────────────────────────────────────────────
# RESET / CLEAR ALL DATA
# ─────────────────────────────────────────────────────────────
#
# Everything the training pipeline produces lives in a handful of
# SHARED, app-wide files/tables — Final_Merged_Student_Data.csv,
# Final_LongForm_Student_Grades.csv, the model_datasets/*.csv exports,
# every trained .pkl in ML_MODEL_DIR, training_state.json, and the
# UploadedDataset rows — none of it is scoped per user. That's why a
# fresh account can see populated charts / a trained Model Eval card
# despite never having uploaded anything themselves: whoever uploaded
# first left data sitting in these shared files, and everyone reads
# the same copy. This route wipes all of it back to a clean slate.

def _delete_files_in(dir_path: str, patterns=('*',)) -> list[str]:
    """Delete files (not subdirectories) directly inside dir_path matching
    any of `patterns`. Returns the list of deleted paths. Safe no-op if
    dir_path doesn't exist."""
    deleted = []
    if not dir_path or not os.path.isdir(dir_path):
        return deleted
    for pattern in patterns:
        for fpath in glob.glob(os.path.join(dir_path, pattern)):
            if os.path.isfile(fpath):
                try:
                    os.remove(fpath)
                    deleted.append(fpath)
                except OSError as e:
                    print(f"[upload_routes] reset: failed to delete {fpath}: {e}")
    return deleted


@upload_bp.route('/api/reset-all-data', methods=['POST'])
def api_reset_all_data():
    """
    Wipes the shared master dataset, every trained model, and the
    upload history — back to the exact 'no model has been trained yet'
    state a brand-new install starts in. This is destructive and
    affects EVERY user of the app, not just the caller, since none of
    the underlying storage is per-user.

    Requires:
      - an authenticated user whose role is in UPLOAD_ALLOWED_ROLES
        (same permission gate as uploading — tighten this to a
        dedicated 'admin' role if/when one exists)
      - a JSON body of {"confirm": "RESET"} as a deliberate
        speed-bump against an accidental call (e.g. a stray button
        click or a retried request)
    """
    user = _current_user()
    if not user:
        return jsonify({'ok': False, 'error': 'Not logged in.'}), 401
    if user.role not in UPLOAD_ALLOWED_ROLES:
        return jsonify({'ok': False, 'error': 'Your role is not permitted to reset data.'}), 403

    body = request.get_json(silent=True) or {}
    if body.get('confirm') != 'RESET':
        return jsonify({
            'ok': False,
            'error': 'Destructive action. Resend with JSON body {"confirm": "RESET"} to proceed.',
        }), 400

    deleted = {
        'model_artifacts'   : _delete_files_in(ML_MODEL_DIR, ('*.pkl', '*.json')),
        'model_datasets'    : _delete_files_in(MODEL_DATASETS_DIR, ('*.csv',)),
        'master_csvs'       : [],
        'raw_uploads'       : _delete_files_in(UNPROCESSED_DATASETS_DIR, ('*',)),
        'processed_csvs'    : _delete_files_in(PROCESSED_DATASETS_DIR, ('*.csv',)),
    }

    # FINAL_MERGED_CSV and the long-form CSV live in PROCESSED_DATASETS_DIR
    # too, so the glob above already caught them — but call out explicitly
    # for the response in case that path ever changes.
    long_form_csv = os.path.join(PROCESSED_DATASETS_DIR, 'Final_LongForm_Student_Grades.csv')
    for p in (FINAL_MERGED_CSV, long_form_csv):
        if p not in deleted['processed_csvs'] and os.path.exists(p):
            try:
                os.remove(p)
                deleted['master_csvs'].append(p)
            except OSError as e:
                print(f"[upload_routes] reset: failed to delete {p}: {e}")

    # Clear upload history (DB) so the audit trail matches reality.
    deleted_records = UploadedDataset.query.delete()
    db.session.commit()

    # Hot-reload the in-memory ML models so dashboards/model-eval stop
    # serving stale predictions from before the reset, same mechanism
    # used after a normal upload finishes training.
    try:
        from ml_route.ml_analysis import reload_models
        reload_models()
    except Exception as _re:
        print(f"[upload_routes] reset: reload_models warning: {_re}")

    total_files_deleted = sum(len(v) for v in deleted.values() if isinstance(v, list))

    return jsonify({
        'ok': True,
        'message': (
            f"Reset complete. Deleted {total_files_deleted} file(s) and "
            f"{deleted_records} upload record(s). The dashboard and Model "
            "Eval will show 'no data' until the next upload."
        ),
        'deleted': deleted,
        'deleted_records': deleted_records,
    })